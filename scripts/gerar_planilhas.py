#!/usr/bin/env python3
"""
Gerador de Planilhas de QA - Trust & Safety Jusbrasil
Gera as 3 planilhas (Planilha Mãe, Mari, Lud) para moderação pelo agente de IA.

Uso:
  python gerar_planilhas.py <arquivo_processos> <tema> [--link-mari URL] [--link-lud URL] [--link-mae URL] [--output-dir ./]

Temas disponíveis: familia, saude, criminal, eca, trabalhista, segredo
"""

import sys
import os
import csv
import argparse
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURAÇÃO POR TEMA
# ─────────────────────────────────────────────

TEMAS = {
    "familia": {
        "nome": "Direito de Família",
        "pergunta_d": "O processo é de Direito de Família?",
        "sim": "SIM, é um processo de direito de familia",
        "nao": "NÃO é um processo de direito de familia",
        "header_agente": "O PROCESSO É DE DIREITO DE FAMÍLIA?",
        "formula_desempenho_mod": (
            '=SE(OU('
            'E(C{r}="SIM, é um processo de direito de familia";REGEXMATCH(D{r};"(?i)por isso o reporte é aprovado|aprovado"));'
            'E(C{r}="NÃO é um processo de direito de familia";REGEXMATCH(D{r};"(?i)reprovado"))'
            ');"✅ ACERTO";"❌ ERRO")'
        ),
    },
    "saude": {
        "nome": "Saúde",
        "pergunta_d": "O processo trata sobre questões de saúde?",
        "sim": "SIM, o processo trata sobre questões de saúde",
        "nao": "NÃO é um processo que trata sobre questões de saúde",
        "header_agente": "O PROCESSO TRATA SOBRE QUESTÕES DE SAÚDE?",
        "formula_desempenho_mod": (
            '=SE(OU('
            'E(C{r}="SIM, o processo trata sobre questões de saúde";D{r}="aprovado");'
            'E(C{r}="NÃO é um processo que trata sobre questões de saúde";D{r}="reprovado")'
            ');"✅ ACERTO";"❌ ERRO")'
        ),
    },
    "criminal": {
        "nome": "Direito Criminal",
        "pergunta_d": "O processo é de Direito Criminal Finalizado?",
        "sim": "SIM, é um processo de direito criminal finalizado",
        "nao": "NÃO é um processo de direito criminal finalizado",
        "header_agente": "O PROCESSO É DE DIREITO CRIMINAL FINALIZADO?",
        "formula_desempenho_mod": (
            '=SE(OU('
            'E(C{r}="SIM, é um processo de direito criminal finalizado";REGEXMATCH(D{r};"(?i)aprovado|por isso o reporte é aprovado"));'
            'E(C{r}="NÃO é um processo de direito criminal finalizado";REGEXMATCH(D{r};"(?i)reprovado"))'
            ');"✅ ACERTO";"❌ ERRO")'
        ),
    },
    "eca": {
        "nome": "ECA / Menor Infrator",
        "pergunta_d": "O processo envolve ECA ou menor infrator?",
        "sim": "SIM, é um processo que envolve ECA ou menor infrator",
        "nao": "NÃO é um processo que envolve ECA ou menor infrator",
        "header_agente": "O PROCESSO ENVOLVE ECA OU MENOR INFRATOR?",
        "formula_desempenho_mod": (
            '=SE(OU('
            'E(C{r}="SIM, é um processo que envolve ECA ou menor infrator";REGEXMATCH(D{r};"(?i)aprovado"));'
            'E(C{r}="NÃO é um processo que envolve ECA ou menor infrator";REGEXMATCH(D{r};"(?i)reprovado"))'
            ');"✅ ACERTO";"❌ ERRO")'
        ),
    },
    "trabalhista": {
        "nome": "Direito Trabalhista",
        "pergunta_d": "O processo é de Direito Trabalhista?",
        "sim": "SIM, é um processo de direito trabalhista",
        "nao": "NÃO é um processo de direito trabalhista",
        "header_agente": "O PROCESSO É DE DIREITO TRABALHISTA?",
        "formula_desempenho_mod": (
            '=SE(OU('
            'E(C{r}="SIM, é um processo de direito trabalhista";REGEXMATCH(D{r};"(?i)aprovado"));'
            'E(C{r}="NÃO é um processo de direito trabalhista";REGEXMATCH(D{r};"(?i)reprovado"))'
            ');"✅ ACERTO";"❌ ERRO")'
        ),
    },
    "segredo": {
        "nome": "Segredo de Justiça",
        "pergunta_d": "O processo está em segredo de justiça?",
        "sim": "SIM, é um processo em segredo de justiça",
        "nao": "NÃO é um processo em segredo de justiça",
        "header_agente": "O PROCESSO ESTÁ EM SEGREDO DE JUSTIÇA?",
        "formula_desempenho_mod": (
            '=SE(OU('
            'E(C{r}="SIM, é um processo em segredo de justiça";REGEXMATCH(D{r};"(?i)aprovado"));'
            'E(C{r}="NÃO é um processo em segredo de justiça";REGEXMATCH(D{r};"(?i)reprovado"))'
            ');"✅ ACERTO";"❌ ERRO")'
        ),
    },
}

# Fórmula de classificação temática (coluna H da aba de desempenho)
# Mesma para todos os temas — compara F (classificação OQJ) com G (classificação agente)
FORMULA_DESEMPENHO_CLASS = (
    '=SE(F{r}="Pendente";"Desconsiderado";SE('
    'OU('
    'E(CONT.SE(F{r};"*SAÚDE*");REGEXMATCH(MINÚSCULA(G{r});"doen[cç]a|erro m[eé]dico|interdi[cç][aã]o|curatela|estigmatizante|cid|psiqu|esquizof|hiv|aids|depress|autismo|tea|borderline|ansiedade|mental|droga|depend[êe]ncia|interna[cç][aã]o|hospitalar|m[eé]dic|cir[uú]rg|prontu|tratamento|medicamento|sa[uú]de|exame|incapacidade|aux[ií]lio-doen[cç]a|invalidez|bpc|loas|neoplas|c[aâ]ncer|tumor"));'
    'E(F{r}="TRABALHISTA";REGEXMATCH(MINÚSCULA(G{r});"trabalhista|trabalho|trt|verbas|fgts|horas extras|rescis[aã]o|clt|adicional|periculosidade|insalubridade|v[ií]nculo|empreg"));'
    'E(CONT.SE(F{r};"*PENAL*");REGEXMATCH(MINÚSCULA(G{r});"criminal|penal|furto|recepta[cç][aã]o|execu[cç][aã]o|pena|crimes|crime|inj[uú]ria|contraven[cç][aã]o|absolvi[cç][aã]o|extin[cç][aã]o|punibilidade|prescri[cç][aã]o|transitado|julgado|maria da penha|viol[êe]ncia dom[eé]stica|amea[çc]a|les[aã]o corporal|medida[s]? protetiva[s]?|inqu[eé]rito|delegacia|pris[aã]o|den[uú]ncia|sursis"));'
    'E(CONT.SE(F{r};"*FAMÍLIA*");REGEXMATCH(MINÚSCULA(G{r});"fam[ií]lia|pensa[oõ]|ado[cç][aã]o|guarda|casamento|alimentos|uni[aã]o est[aá]vel|div[oó]rcio|partilha|paternidade|filia[cç][aã]o"));'
    'E(CONT.SE(F{r};"*CRIMES SEXUAIS*");REGEXMATCH(MINÚSCULA(G{r});"sexual|[oó]dio|liberdade pessoal|honra|estupro|importuna[cç][aã]o|ass[eé]dio|vulner[aá]vel"));'
    'E(CONT.SE(F{r};"*MENOR*");REGEXMATCH(MINÚSCULA(G{r});"menor|eca|infrator|adolescente|crian[cç]a|ato infracional"));'
    'E(F{r}="SEGREDO DE JUSTIÇA";REGEXMALCH(MINÚSCULA(G{r});"segredo|sigilo|sigiloso|reprovado|n/a"))'
    ');"✅ ACERTO";"❌ ERRO"))'
)

# ─────────────────────────────────────────────
# OPÇÕES FIXAS (iguais para todos os temas)
# ─────────────────────────────────────────────

OPCOES_E = [
    "SIM, é um caso de PUBLICIDADE PROCESSUAL ou de RESTRIÇÃO A IDENTIFICAÇÃO DA PARTE",
    "SIM, é um caso de DADO PESSOAL EXPOSTO",
    "SIM, é um caso de ENVOLVIMENTO INCORRETO",
    "NÃO se enquadra em nenhum critério e deveria ser REPROVADO",
]

OPCOES_F = [
    "FAMÍLIA: Filiação, Adoção e Guarda / Casamento e União / Pensão Alimentícia",
    "CRIMES SEXUAIS OU DE ÓDIO: Crime Sexual / Vítima ou testemunha de crime / Crime de Ódio",
    "QUESTÕES DE SAÚDE: Doença Estigmatizante / Erro Médico / Procedimento Médico / Doença Geral",
    "TRABALHISTA",
    "MENOR: Menor Infrator / ECA",
    "PENAL: Criminal Finalizado / Violência Doméstica",
    "SEGREDO DE JUSTIÇA",
]

# ─────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────

COR_AZUL_ESCURO = "1F3864"   # Cabeçalho grupo nível 1
COR_AZUL_MEDIO = "2E5B9C"   # Cabeçalho grupo nível 2
COR_AZUL_CLARO = "BDD7EE"   # Sub-cabeçalho Mari/Lud
COR_VERDE      = "E2EFDA"   # Desempate
COR_CINZA      = "F2F2F2"   # Linhas alternadas

def estilo_cabecalho(cor_fundo, negrito=True, tamanho=10, cor_fonte="FFFFFF", wrap=True):
    font = Font(bold=negrito, size=tamanho, color=cor_fonte)
    fill = PatternFill(fill_type="solid", fgColor=cor_fundo)
    align = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return font, fill, align

def borda_fina():
    lado = Side(style="thin")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def aplicar_celula(ws, row, col, valor, cor_fundo=None, negrito=False, cor_fonte="000000",
                   wrap=True, centralizar=True, formula=False):
    cell = ws.cell(row=row, column=col, value=valor)
    if cor_fundo:
        font_color = "FFFFFF" if cor_fundo in [COR_AZUL_ESCURO, COR_AZUL_MEDIO] else cor_fonte
        cell.font = Font(bold=negrito, size=10, color=font_color)
        cell.fill = PatternFill(fill_type="solid", fgColor=cor_fundo)
    else:
        cell.font = Font(bold=negrito, size=10, color=cor_fonte)
    cell.alignment = Alignment(
        horizontal="center" if centralizar else "left",
        vertical="center",
        wrap_text=wrap
    )
    cell.border = borda_fina()
    return cell

# ─────────────────────────────────────────────
# LEITURA DOS DADOS DE ENTRADA
# ─────────────────────────────────────────────

def ler_processos(arquivo):
    """
    Lê arquivo de entrada (CSV ou XLSX) e retorna lista de dicts:
    [{"id": 1, "link": "https://...", "numero": "0001234-56.2024..."}]
    """
    ext = os.path.splitext(arquivo)[1].lower()
    processos = []

    if ext == ".csv":
        with open(arquivo, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader, 1):
                # Tenta inferir colunas por nome ou posição
                link = (row.get("link") or row.get("Link") or row.get("LINK") or
                        row.get("url") or row.get("URL") or "")
                numero = (row.get("numero") or row.get("Número do processo") or
                          row.get("numero_processo") or row.get("NÚMERO") or "")
                processos.append({"id": i, "link": link.strip(), "numero": numero.strip()})

    elif ext in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(arquivo, data_only=True)
        ws = wb.active

        # Detecta linha de cabeçalho (primeira linha com "link" ou "processo")
        header_row = 1
        headers = {}
        for row in ws.iter_rows(min_row=1, max_row=5):
            for cell in row:
                val = str(cell.value or "").lower()
                if "link" in val:
                    headers["link"] = cell.column
                    header_row = cell.row
                if "número" in val or "numero" in val or "processo" in val:
                    headers["numero"] = cell.column
                    header_row = cell.row

        idx = 1
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(row):
                continue
            link = str(row[headers.get("link", 2) - 1] or "").strip()
            numero = str(row[headers.get("numero", 3) - 1] or "").strip()
            if link or numero:
                processos.append({"id": idx, "link": link, "numero": numero})
                idx += 1
    else:
        raise ValueError(f"Formato não suportado: {ext}. Use .csv ou .xlsx")

    return processos

# ─────────────────────────────────────────────
# GERAÇÃO: PLANILHA EXECUÇÃO (MARI / LUD)
# ─────────────────────────────────────────────

def gerar_planilha_execucao(processos, tema_cfg, nome_pessoa, link_saida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Execução"
    ws.sheet_view.showGridLines = True

    # Larguras das colunas
    larguras = {"A": 8, "B": 55, "C": 28, "D": 28, "E": 45, "F": 55, "G": 35}
    for col, larg in larguras.items():
        ws.column_dimensions[col].width = larg
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 50
    ws.row_dimensions[3].height = 16

    # ── Linha 1: Grupos ──
    ws.merge_cells("A1:C1")
    aplicar_celula(ws, 1, 1, "DADOS DO PROCESSO", COR_AZUL_ESCURO, negrito=True)
    ws.merge_cells("D1:G1")
    aplicar_celula(ws, 1, 4, "EXECUÇÃO", COR_AZUL_ESCURO, negrito=True)

    # ── Linha 2: Cabeçalhos ──
    cabecalhos = [
        (1, "ID"), (2, "Link do Processo"), (3, "Número do processo"),
        (4, tema_cfg["pergunta_d"]),
        (5, "O processo se enquadra nos critérios institucionais de aprovação?"),
        (6, "Sendo caso de PUBLICIDADE PROCESSUAL ou de RESTRIÇÃO A IDENTIFICAÇÃO DA PARTE, em qual tema o processo se enquadra?"),
        (7, "OBSERVAÇÕES"),
    ]
    for col, titulo in cabecalhos:
        aplicar_celula(ws, 2, col, titulo, COR_AZUL_MEDIO, negrito=True)

    # ── Linha 3: Sub-cabeçalho (pessoa executora) ──
    for col in range(1, 8):
        aplicar_celula(ws, 3, col, nome_pessoa if col >= 4 else "", COR_AZUL_CLARO, negrito=False, cor_fonte="000000")

    # ── Linhas de dados ──
    for processo in processos:
        r = processo["id"] + 3  # dados começam na linha 4
        aplicar_celula(ws, r, 1, processo["id"], centralizar=True)
        aplicar_celula(ws, r, 2, processo["link"], centralizar=False)
        aplicar_celula(ws, r, 3, processo["numero"], centralizar=False)
        # Colunas D, E, F, G vazias — preenchidas manualmente
        for col in range(4, 8):
            aplicar_celula(ws, r, col, "", centralizar=True)

        # Linhas alternadas
        if processo["id"] % 2 == 0:
            for col in range(1, 8):
                cell = ws.cell(row=r, column=col)
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = PatternFill(fill_type="solid", fgColor=COR_CINZA)

    wb.save(link_saida)
    return link_saida

# ─────────────────────────────────────────────
# GERAÇÃO: PLANILHA MÃE
# ─────────────────────────────────────────────

def gerar_planilha_mae(processos, tema_cfg, link_mari, link_lud, link_mae, link_saida):
    wb = openpyxl.Workbook()

    # ── ABA 1: Execução ──
    ws_exec = wb.active
    ws_exec.title = "Execução"

    n = len(processos)
    linha_fim = n + 3  # dados de 4 até 4+n-1

    # Larguras
    for col_idx, larg in enumerate([8, 55, 28, 25, 25, 35, 35, 35, 35, 30, 30, 28, 35, 35], 1):
        ws_exec.column_dimensions[get_column_letter(col_idx)].width = larg
    ws_exec.row_dimensions[1].height = 20
    ws_exec.row_dimensions[2].height = 50
    ws_exec.row_dimensions[3].height = 16

    # Linha 1: Grupos
    ws_exec.merge_cells("A1:C1")
    aplicar_celula(ws_exec, 1, 1, "DADOS DO PROCESSO", COR_AZUL_ESCURO, negrito=True)
    ws_exec.merge_cells("D1:K1")
    aplicar_celula(ws_exec, 1, 4, "EXECUÇÃO", COR_AZUL_ESCURO, negrito=True)
    ws_exec.merge_cells("L1:N1")
    aplicar_celula(ws_exec, 1, 12, "DESEMPATE", COR_VERDE, negrito=True, cor_fonte="000000")

    # Linha 2: Cabeçalhos
    pergunta = tema_cfg["pergunta_d"]
    cabecalhos_exec = [
        (1, "ID"), (2, "Link do Processo"), (3, "Número do processo"),
        (4, f"{pergunta} — Mari"), (5, f"{pergunta} — Lud"),
        (6, "O processo se enquadra nos critérios institucionais de aprovação? — Mari"),
        (7, "O processo se enquadra nos critérios institucionais de aprovação? — Lud"),
        (8, "Sendo caso de PUBLICIDADE PROCESSUAL, em qual tema? — Mari"),
        (9, "Sendo caso de PUBLICIDADE PROCESSUAL, em qual tema? — Lud"),
        (10, "Comentários — Mari"), (11, "Comentários — Lud"),
        (12, f"DESEMPATE — {pergunta}"),
        (13, "DESEMPATE — Critérios de aprovação"),
        (14, "DESEMPATE — Tema"),
    ]
    for col, titulo in cabecalhos_exec:
        cor = COR_VERDE if col >= 12 else COR_AZUL_MEDIO
        fonte = "000000" if col >= 12 else "FFFFFF"
        aplicar_celula(ws_exec, 2, col, titulo, cor, negrito=True, cor_fonte=fonte)

    # Linha 3: Sub-cabeçalhos Mari/Lud
    for col in range(1, 15):
        if col in (4, 6, 8, 10):
            val = "Mari"
        elif col in (5, 7, 9, 11):
            val = "Lud"
        else:
            val = ""
        aplicar_celula(ws_exec, 3, col, val, COR_AZUL_CLARO, cor_fonte="000000")

    # IMPORTRANGE — fórmulas em bloco (célula D4)
    # Google Sheets: =IMPORTRANGE aceita range expandido
    range_fim = n + 3  # ex: D4:D103 para 100 processos

    importranges = {
        4:  f'=IMPORTRANGE("{link_mari}";"Execução!D4:D{range_fim}")',
        5:  f'=IMPORTRANGE("{link_lud}";"Execução!D4:D{range_fim}")',
        6:  f'=IMPORTRANGE("{link_mari}";"Execução!E4:E{range_fim}")',
        7:  f'=IMPORTRANGE("{link_lud}";"Execução!E4:E{range_fim}")',
        8:  f'=IMPORTRANGE("{link_mari}";"Execução!F4:F{range_fim}")',
        9:  f'=IMPORTRANGE("{link_lud}";"Execução!F4:F{range_fim}")',
        10: f'=IMPORTRANGE("{link_mari}";"Execução!G4:G{range_fim}")',
        11: f'=IMPORTRANGE("{link_lud}";"Execução!G4:G{range_fim}")',
    }

    # Linha 4: IMPORTRANGE (fórmula em bloco — só primeira linha; Google Sheets expande)
    for col, formula in importranges.items():
        cell = ws_exec.cell(row=4, column=col, value=formula)
        cell.alignment = Alignment(wrap_text=True)
        cell.border = borda_fina()

    # Dados do processo + desempate linha a linha
    for processo in processos:
        r = processo["id"] + 3
        aplicar_celula(ws_exec, r, 1, processo["id"], centralizar=True)
        aplicar_celula(ws_exec, r, 2, processo["link"], centralizar=False)
        aplicar_celula(ws_exec, r, 3, processo["numero"], centralizar=False)

        # Desempate L, M, N
        formula_l = f'=SE(OU(D{r}="";E{r}="");"Pendente";SE(D{r}=E{r};D{r};"DESEMPATAR"))'
        formula_m = f'=SE(OU(F{r}="";G{r}="");"Pendente";SE(F{r}=G{r};F{r};"DESEMPATAR"))'
        formula_n = f'=SE(OU(H{r}="";I{r}="");"Pendente";SE(H{r}=I{r};H{r};"DESEMPATAR"))'

        for col, formula in [(12, formula_l), (13, formula_m), (14, formula_n)]:
            cell = ws_exec.cell(row=r, column=col, value=formula)
            cell.fill = PatternFill(fill_type="solid", fgColor=COR_VERDE)
            cell.border = borda_fina()
            cell.alignment = Alignment(wrap_text=True)

    # ── ABA 2: Desempenho do Agente ──
    ws_desemp = wb.create_sheet("Desempenho do Agente")

    for col_idx, larg in enumerate([8, 55, 28, 45, 22, 45, 45, 22, 20, 35], 1):
        ws_desemp.column_dimensions[get_column_letter(col_idx)].width = larg
    ws_desemp.row_dimensions[1].height = 20
    ws_desemp.row_dimensions[2].height = 40

    # Linha 1: Grupos
    ws_desemp.merge_cells("A1:B1")
    aplicar_celula(ws_desemp, 1, 1, "", COR_AZUL_ESCURO)
    ws_desemp.merge_cells("C1:E1")
    aplicar_celula(ws_desemp, 1, 3, tema_cfg["header_agente"], COR_AZUL_ESCURO, negrito=True)
    ws_desemp.merge_cells("F1:H1")
    aplicar_celula(ws_desemp, 1, 6, "EM QUAL TEMA O PROCESSO SE ENQUADRA?", COR_AZUL_ESCURO, negrito=True)
    aplicar_celula(ws_desemp, 1, 10, "COMENTÁRIOS", COR_AZUL_ESCURO, negrito=True)

    # Linha 2: Cabeçalhos
    cabs_desemp = [
        (1, "ID"), (2, "LINK"),
        (3, "MODERAÇÃO OQJ"), (4, "MODERAÇÃO AGENTE (EXPANDIDO)"), (5, "DESEMPENHO DO AGENTE"),
        (6, "CLASSIFICAÇÃO OQJ"), (7, "CLASSIFICAÇÃO AGENTE (EXPANDIDO)"), (8, "DESEMPENHO DO AGENTE"),
        (10, "COMENTÁRIOS"),
    ]
    for col, titulo in cabs_desemp:
        aplicar_celula(ws_desemp, 2, col, titulo, COR_AZUL_MEDIO, negrito=True)
    # col 9 vazia
    aplicar_celula(ws_desemp, 2, 9, "", COR_AZUL_MEDIO)

    # IMPORTRANGE para colunas C e F (desempate da aba Execução)
    range_desemp_fim = n + 2  # dados de 3 até 3+n-1 na aba desempenho; em Execução são L4:L{range_fim}
    cell_c = ws_desemp.cell(row=3, column=3,
        value=f'=IMPORTRANGE("{link_mae}";"Execução!L4:L{range_fim}")')
    cell_c.alignment = Alignment(wrap_text=True)
    cell_c.border = borda_fina()

    cell_f = ws_desemp.cell(row=3, column=6,
        value=f'=IMPORTRANGE("{link_mae}";"Execução!N4:N{range_fim}")')
    cell_f.alignment = Alignment(wrap_text=True)
    cell_f.border = borda_fina()

    # Dados + fórmulas de desempenho linha a linha
    for processo in processos:
        r = processo["id"] + 2  # linha 3 = ID 1
        aplicar_celula(ws_desemp, r, 1, processo["id"], centralizar=True)
        aplicar_celula(ws_desemp, r, 2, processo["link"], centralizar=False)
        # Col D e G: fornecidas pelo usuário (respostas do agente) — deixar vazio
        for col in [4, 7]:
            aplicar_celula(ws_desemp, r, col, "", centralizar=False)
        # Col 9 e 10 vazias
        for col in [9, 10]:
            aplicar_celula(ws_desemp, r, col, "", centralizar=False)

        # Col E: Desempenho de moderação
        formula_e = tema_cfg["formula_desempenho_mod"].format(r=r)
        cell_e = ws_desemp.cell(row=r, column=5, value=formula_e)
        cell_e.border = borda_fina()
        cell_e.alignment = Alignment(horizontal="center", vertical="center")

        # Col H: Desempenho de classificação
        formula_h = FORMULA_DESEMPENHO_CLASS.format(r=r)
        cell_h = ws_desemp.cell(row=r, column=8, value=formula_h)
        cell_h.border = borda_fina()
        cell_h.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(link_saida)
    return link_saida

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Gerador de Planilhas de QA - Jusbrasil T&S")
    parser.add_argument("arquivo", help="Arquivo com os processos (.csv ou .xlsx)")
    parser.add_argument("tema", choices=list(TEMAS.keys()),
                        help="Tema: familia, saude, criminal, eca, trabalhista, segredo")
    parser.add_argument("--link-mari", default="LINK-PLANILHA-MARI",
                        help="URL da planilha Mari no Google Sheets")
    parser.add_argument("--link-lud", default="LINK-PLANILHA-LUD",
                        help="URL da planilha Lud no Google Sheets")
    parser.add_argument("--link-mae", default="LINK-PLANILHA-MAE",
                        help="URL da planilha Mãe no Google Sheets")
    parser.add_argument("--output-dir", default=".",
                        help="Diretório de saída (padrão: diretório atual)")
    args = parser.parse_args()

    tema_cfg = TEMAS[args.tema]
    data_hoje = datetime.today().strftime("%Y.%m.%d")
    nome_tema = tema_cfg["nome"].replace(" ", "_").replace("/", "")

    print(f"📂 Lendo processos de: {args.arquivo}")
    processos = ler_processos(args.arquivo)
    print(f"✅ {len(processos)} processos carregados")

    os.makedirs(args.output_dir, exist_ok=True)

    # Caminhos de saída
    path_mari = os.path.join(args.output_dir, f"[EXECUÇÃO MARI] {data_hoje} - QA - {nome_tema}.xlsx")
    path_lud  = os.path.join(args.output_dir, f"[EXECUÇÃO LUD] {data_hoje} - QA - {nome_tema}.xlsx")
    path_mae  = os.path.join(args.output_dir, f"[PLANILHA MÃE] {data_hoje} - QA - {nome_tema}.xlsx")

    print("📝 Gerando planilha Mari...")
    gerar_planilha_execucao(processos, tema_cfg, "Mariana", path_mari)
    print(f"   ✅ {path_mari}")

    print("📝 Gerando planilha Lud...")
    gerar_planilha_execucao(processos, tema_cfg, "Ludmila", path_lud)
    print(f"   ✅ {path_lud}")

    print("📝 Gerando planilha Mãe...")
    gerar_planilha_mae(processos, tema_cfg, args.link_mari, args.link_lud, args.link_mae, path_mae)
    print(f"   ✅ {path_mae}")

    print("\n🎉 Planilhas geradas com sucesso!")
    print("\n📌 Próximos passos:")
    print("  1. Faça upload das 3 planilhas no Google Sheets")
    print("  2. Copie as URLs das planilhas Mari e Lud")
    print("  3. Na planilha Mãe, substitua LINK-PLANILHA-MARI e LINK-PLANILHA-LUD pelas URLs reais")
    print("  4. Autorize o IMPORTRANGE quando solicitado pelo Google Sheets")
    if args.link_mae == "LINK-PLANILHA-MAE":
        print("  5. Na aba Desempenho do Agente, substitua LINK-PLANILHA-MAE pela URL da planilha Mãe")

if __name__ == "__main__":
    main()
