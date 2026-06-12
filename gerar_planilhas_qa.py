"""
GERADOR AUTOMÁTICO DE PLANILHAS QA - TRUST & SAFETY (JUSBRASIL)
================================================================
Uso:
    python3 gerar_planilhas_qa.py

O script lê um arquivo de entrada com os casos (.xlsx ou .csv) e gera:
  1. Planilha de execução individual para Ludmila
  2. Planilha de execução individual para Mariana
  3. Planilha Mãe (consolidação com todas as abas)

Formato esperado do arquivo de entrada (colunas obrigatórias):
  - report_id        → ID do reporte (ex: ffd52ef7-automatic:...)
  - lawsuit_id       → ID do processo (ex: 4b0b9f72-...)
  - link             → URL do processo no Jusbrasil
  - numero_processo  → Número do processo (ex: 6037209-42.2025.8.09.0051)
  - nome_parte       → Nome da parte
  - cpf              → CPF da parte
"""

import os
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# PALETA DE CORES (extraída dos modelos reais)
# ─────────────────────────────────────────────
COR_VERDE_ESCURO  = "1C664F"   # DADOS DO PROCESSO / DA PARTE
COR_VERDE_CLARO   = "6AA84F"   # EXECUÇÃO
COR_AZUL_CLARO    = "9FC5E8"   # Mari
COR_ROXO_CLARO    = "B4A7D6"   # Lud
COR_VERMELHO      = "85200C"   # DESEMPATE
COR_BRANCO        = "FFFFFF"
COR_LINK          = "1155CC"   # Azul hiperlink

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=COR_BRANCO, size=10, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    side = Side(style='thin', color='000000')
    return Border(left=side, right=side, top=side, bottom=side)

def estilo_header_secao(cor_bg):
    """Linha 1: headers de seção (ex: DADOS DO PROCESSO, EXECUÇÃO)"""
    return {
        "fill": fill(cor_bg),
        "font": font(bold=True, color=COR_BRANCO, size=10),
        "alignment": align(),
        "border": thin_border()
    }

def estilo_header_coluna(cor_bg):
    """Linha 2: headers de coluna (ex: Link do Processo, CPF)"""
    return {
        "fill": fill(cor_bg),
        "font": font(bold=True, color=COR_BRANCO, size=9),
        "alignment": align(wrap=True),
        "border": thin_border()
    }

def estilo_header_analista(cor_bg):
    """Linha 3 (Planilha Mãe): nome do analista"""
    return {
        "fill": fill(cor_bg),
        "font": font(bold=True, color=COR_BRANCO, size=9),
        "alignment": align(),
        "border": thin_border()
    }

def estilo_dado():
    return {
        "font": Font(size=9, name="Calibri"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": thin_border()
    }

def estilo_link():
    return {
        "font": Font(size=9, color=COR_LINK, name="Calibri"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": thin_border()
    }

def aplicar_estilo(cell, estilo):
    for attr, val in estilo.items():
        setattr(cell, attr, val)

def mesclar_e_estilizar(ws, start_row, start_col, end_col, valor, estilo):
    col_ini = get_column_letter(start_col)
    col_fim = get_column_letter(end_col)
    if start_col != end_col:
        ws.merge_cells(f"{col_ini}{start_row}:{col_fim}{start_row}")
    cell = ws.cell(start_row, start_col, valor)
    aplicar_estilo(cell, estilo)

def estilizar_coluna_vazia(ws, row, start_col, end_col, estilo):
    """Preenche células mescladas de apoio com estilo (sem valor)"""
    for col in range(start_col, end_col + 1):
        c = ws.cell(row, col)
        aplicar_estilo(c, estilo)

# ─────────────────────────────────────────────
# PERGUNTAS POR AGENTE (configurável)
# ─────────────────────────────────────────────
PERGUNTAS = {
    "familia": {
        "q1": "O processo trata de Direito de Família?",
        "q2": "O processo se enquadra nos critérios institucionais de aprovação?",
        "q3": "Sendo caso de PUBLICIDADE PROCESSUAL ou de RESTRIÇÃO A IDENTIFICAÇÃO DA PARTE, em qual tema o processo se enquadra?",
    },
    "saude": {
        "q1": "O processo trata sobre questões de saúde?",
        "q2": "O processo se enquadra nos critérios institucionais de aprovação?",
        "q3": "Sendo caso de PUBLICIDADE PROCESSUAL ou de RESTRIÇÃO A IDENTIFICAÇÃO DA PARTE, em qual tema o processo se enquadra?",
    },
    "custom": None  # preenchido pelo usuário
}

# ─────────────────────────────────────────────
# LEITURA DO ARQUIVO DE ENTRADA
# ─────────────────────────────────────────────
COLUNAS_ENTRADA = {
    "report_id":       ["report_id", "Report ID", "ID", "id reporte", "reporte_id"],
    "lawsuit_id":      ["lawsuit_id", "Lawsuit ID", "id processo", "processo_id"],
    "link":            ["link", "Link do Processo", "url", "URL"],
    "numero_processo": ["numero_processo", "Número do processo", "numero processo", "process_number"],
    "nome_parte":      ["nome_parte", "Nome da parte", "nome", "name"],
    "cpf":             ["cpf", "CPF", "cpf_parte"],
}

def normalizar(texto):
    return str(texto).strip().lower().replace(" ", "_") if texto else ""

def encontrar_coluna(headers, candidatos):
    """Retorna o índice (0-based) da coluna no header que bate com algum candidato"""
    for i, h in enumerate(headers):
        norm = normalizar(h)
        for c in candidatos:
            if normalizar(c) == norm or normalizar(c) in norm:
                return i
    return None

def ler_arquivo_entrada(caminho):
    """Lê CSV ou XLSX e retorna lista de dicts com os dados dos casos"""
    ext = os.path.splitext(caminho)[1].lower()
    rows = []

    if ext == ".csv":
        import csv
        with open(caminho, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader)
            all_rows = list(reader)
    elif ext in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(caminho, data_only=True)
        ws = wb.active
        all_data = list(ws.values)
        headers = [str(h) if h is not None else "" for h in all_data[0]]
        all_rows = [list(r) for r in all_data[1:]]
    else:
        raise ValueError(f"Formato não suportado: {ext}. Use .xlsx ou .csv")

    # Mapear colunas automaticamente
    idx = {}
    for campo, candidatos in COLUNAS_ENTRADA.items():
        pos = encontrar_coluna(headers, candidatos)
        if pos is not None:
            idx[campo] = pos

    obrigatorias = ["report_id", "link"]
    for campo in obrigatorias:
        if campo not in idx:
            raise ValueError(
                f"Coluna obrigatória não encontrada: '{campo}'\n"
                f"Headers detectados: {headers}\n"
                f"Renomeie a coluna ou ajuste COLUNAS_ENTRADA no script."
            )

    for row in all_rows:
        if not any(v for v in row if v):
            continue
        caso = {}
        for campo, pos in idx.items():
            caso[campo] = str(row[pos]).strip() if pos < len(row) and row[pos] is not None else ""
        # Fallbacks
        caso.setdefault("lawsuit_id", "")
        caso.setdefault("numero_processo", "")
        caso.setdefault("nome_parte", "")
        caso.setdefault("cpf", "")
        rows.append(caso)

    return rows

# ─────────────────────────────────────────────
# GERADOR: PLANILHA INDIVIDUAL DO ANALISTA
# ─────────────────────────────────────────────
def gerar_planilha_individual(casos, nome_analista, perguntas, nome_agente, data_rodada, pasta_saida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Execução"

    # Altura das linhas de header
    ws.row_dimensions[1].height = 27.75
    ws.row_dimensions[2].height = 27.75

    # Larguras das colunas
    larguras = {
        "A": 12.63, "B": 12.63,
        "C": 21.38, "D": 22.5,
        "E": 30.5,  "F": 11.88,
        "G": 44.38, "H": 39.13, "I": 48.0, "J": 30.0
    }
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    sec_branco  = estilo_header_secao("FFFFFF")
    sec_verde_e = estilo_header_secao(COR_VERDE_ESCURO)
    sec_verde_c = estilo_header_secao(COR_VERDE_CLARO)
    col_verde_e = estilo_header_coluna(COR_VERDE_ESCURO)
    col_verde_c = estilo_header_coluna(COR_VERDE_CLARO)

    # ── Linha 1: seções
    mesclar_e_estilizar(ws, 1, 1, 2, "DADOS DO REPORTE", sec_branco)
    mesclar_e_estilizar(ws, 1, 3, 4, "DADOS DO PROCESSO", sec_verde_e)
    mesclar_e_estilizar(ws, 1, 5, 6, "DADOS DA PARTE", sec_verde_e)
    mesclar_e_estilizar(ws, 1, 7, 10, "EXECUÇÃO", sec_verde_c)
    # Força cor preta no header branco
    ws.cell(1, 1).font = Font(bold=True, size=10, name="Calibri", color="000000")

    # ── Linha 2: colunas
    headers_l2 = [
        (1, "ID", "FFFFFF"),
        (2, "ID", "FFFFFF"),
        (3, "Link do Processo", COR_VERDE_ESCURO),
        (4, "Número do processo", COR_VERDE_ESCURO),
        (5, "Nome da parte", COR_VERDE_ESCURO),
        (6, "CPF", COR_VERDE_ESCURO),
        (7, perguntas["q1"], COR_VERDE_CLARO),
        (8, perguntas["q2"], COR_VERDE_CLARO),
        (9, perguntas["q3"], COR_VERDE_CLARO),
        (10, "OBSERVAÇÕES", COR_VERDE_CLARO),
    ]
    for col, texto, cor in headers_l2:
        cell = ws.cell(2, col, texto)
        aplicar_estilo(cell, estilo_header_coluna(cor))
        if cor == "FFFFFF":
            cell.font = Font(bold=True, size=9, name="Calibri", color="000000")

    # ── Dados
    for i, caso in enumerate(casos, start=3):
        ws.row_dimensions[i].height = 40

        dados_row = [
            caso.get("report_id", ""),
            caso.get("lawsuit_id", ""),
            caso.get("link", ""),
            caso.get("numero_processo", ""),
            caso.get("nome_parte", ""),
            caso.get("cpf", ""),
            "", "", "", ""  # colunas de execução (analista preenche)
        ]
        for col_idx, valor in enumerate(dados_row, 1):
            cell = ws.cell(i, col_idx, valor)
            if col_idx == 3 and valor.startswith("http"):
                aplicar_estilo(cell, estilo_link())
                cell.hyperlink = valor
            else:
                aplicar_estilo(cell, estilo_dado())

    # Congelar painel no cabeçalho
    ws.freeze_panes = "A3"

    # Salvar
    safe_nome = nome_analista.replace(" ", "_")
    filename = f"[EXECUÇÃO {nome_analista.upper()}] {data_rodada} - QA - {nome_agente}.xlsx"
    out = os.path.join(pasta_saida, filename)
    wb.save(out)
    print(f"  ✅ Gerado: {filename}")
    return out

# ─────────────────────────────────────────────
# GERADOR: PLANILHA MÃE
# ─────────────────────────────────────────────
def gerar_planilha_mae(casos, perguntas, nome_agente, data_rodada, pasta_saida):
    wb = openpyxl.Workbook()

    # ── ABA: Execução ──────────────────────────────────────────────────
    ws_exec = wb.active
    ws_exec.title = "Execução"

    ws_exec.row_dimensions[1].height = 27.75
    ws_exec.row_dimensions[2].height = 27.75
    ws_exec.row_dimensions[3].height = 20

    larguras_mae = {
        "A": 12.63, "B": 12.63,
        "C": 21.38, "D": 22.5,
        "E": 30.5,  "F": 11.88,
        "G": 29.75, "H": 26.63,
        "I": 33.0,  "J": 31.88,
        "K": 31.25, "L": 31.63,
        "M": 20.0,  "N": 20.0,
        "O": 41.63, "P": 49.5, "Q": 43.5
    }
    for col, w in larguras_mae.items():
        ws_exec.column_dimensions[col].width = w

    sec_branco  = estilo_header_secao("FFFFFF")
    sec_verde_e = estilo_header_secao(COR_VERDE_ESCURO)
    sec_verde_c = estilo_header_secao(COR_VERDE_CLARO)
    sec_vermelho = estilo_header_secao(COR_VERMELHO)
    col_verde_e  = estilo_header_coluna(COR_VERDE_ESCURO)
    col_verde_c  = estilo_header_coluna(COR_VERDE_CLARO)

    # Linha 1: seções
    mesclar_e_estilizar(ws_exec, 1, 1, 2, "DADOS DO REPORTE", sec_branco)
    ws_exec.cell(1, 1).font = Font(bold=True, size=10, name="Calibri", color="000000")
    mesclar_e_estilizar(ws_exec, 1, 3, 4, "DADOS DO PROCESSO", sec_verde_e)
    mesclar_e_estilizar(ws_exec, 1, 5, 6, "DADOS DA PARTE", sec_verde_e)
    mesclar_e_estilizar(ws_exec, 1, 7, 14, "EXECUÇÃO", sec_verde_c)
    mesclar_e_estilizar(ws_exec, 1, 15, 17, "DESEMPATE", sec_vermelho)

    # Linha 2: colunas
    # IDs sem destaque
    for col in [1, 2]:
        cell = ws_exec.cell(2, col, "")
        aplicar_estilo(cell, estilo_header_coluna("FFFFFF"))
        cell.font = Font(bold=True, size=9, name="Calibri", color="000000")

    for col, texto in [(3, "Link do Processo"), (4, "Número do processo"),
                       (5, "Nome da parte"), (6, "CPF")]:
        aplicar_estilo(ws_exec.cell(2, col, texto), col_verde_e)

    # G2:H2 merged, I2:J2 merged, K2:L2 merged
    mesclar_e_estilizar(ws_exec, 2, 7, 8, perguntas["q1"], col_verde_c)
    mesclar_e_estilizar(ws_exec, 2, 9, 10, perguntas["q2"], col_verde_c)
    mesclar_e_estilizar(ws_exec, 2, 11, 12, perguntas["q3"], col_verde_c)
    aplicar_estilo(ws_exec.cell(2, 13, "Comentários"), col_verde_c)
    aplicar_estilo(ws_exec.cell(2, 14, "Comentários"), col_verde_c)

    # O2:O3, P2:P3, Q2:Q3 (merged vertical para desempate)
    ws_exec.merge_cells("O2:O3")
    ws_exec.merge_cells("P2:P3")
    ws_exec.merge_cells("Q2:Q3")
    for col, texto in [
        (15, perguntas["q1"]),
        (16, perguntas["q2"]),
        (17, perguntas["q3"])
    ]:
        aplicar_estilo(ws_exec.cell(2, col, texto), col_verde_c)

    # Linha 3: IDs + nomes dos analistas
    for col in [1, 2]:
        cell = ws_exec.cell(3, col, "ID" if col == 1 else "ID")
        aplicar_estilo(cell, estilo_dado())

    for col in [3, 4, 5, 6]:
        cell = ws_exec.cell(3, col, "")
        aplicar_estilo(cell, col_verde_e)

    # Mari = azul, Lud = roxo
    cor_mari = COR_AZUL_CLARO
    cor_lud  = COR_ROXO_CLARO
    pares = [(7, "Mari", cor_mari), (8, "Lud", cor_lud),
             (9, "Mari", cor_mari), (10, "Lud", cor_lud),
             (11, "Mari", cor_mari), (12, "Lud", cor_lud),
             (13, "Mari", cor_mari), (14, "Lud", cor_lud)]
    for col, nome, cor in pares:
        aplicar_estilo(ws_exec.cell(3, col, nome), estilo_header_analista(cor))

    # Dados
    for i, caso in enumerate(casos, start=4):
        ws_exec.row_dimensions[i].height = 40
        dados = [
            caso.get("report_id", ""),
            caso.get("lawsuit_id", ""),
            caso.get("link", ""),
            caso.get("numero_processo", ""),
            caso.get("nome_parte", ""),
            caso.get("cpf", ""),
        ]
        for col_idx, valor in enumerate(dados, 1):
            cell = ws_exec.cell(i, col_idx, valor)
            if col_idx == 3 and valor.startswith("http"):
                aplicar_estilo(cell, estilo_link())
                cell.hyperlink = valor
            else:
                aplicar_estilo(cell, estilo_dado())
        # Colunas de execução e desempate (vazias para preenchimento)
        for col in range(7, 18):
            aplicar_estilo(ws_exec.cell(i, col, ""), estilo_dado())

    ws_exec.freeze_panes = "A4"

    # ── ABA: Revisão ───────────────────────────────────────────────────
    ws_rev = wb.create_sheet("Revisão")
    ws_rev.row_dimensions[1].height = 27.75
    ws_rev.row_dimensions[2].height = 27.75

    revisao_cols = {
        "A": 20.0, "B": 21.38, "C": 22.5, "D": 30.5, "E": 11.88,
        "F": 35.0, "G": 35.0, "H": 45.0, "I": 15.0
    }
    for col, w in revisao_cols.items():
        ws_rev.column_dimensions[col].width = w

    mesclar_e_estilizar(ws_rev, 1, 1, 1, "Auditoria final LEx", sec_branco)
    ws_rev.cell(1, 1).font = Font(bold=True, size=10, name="Calibri", color="000000")
    mesclar_e_estilizar(ws_rev, 1, 2, 5, "DADOS DO PROCESSO", sec_verde_e)
    mesclar_e_estilizar(ws_rev, 1, 6, 8, "RESULTADO - DESEMPATE", sec_verde_c)
    aplicar_estilo(ws_rev.cell(1, 9, "REVISADO"), sec_branco)
    ws_rev.cell(1, 9).font = Font(bold=True, size=10, name="Calibri", color="000000")

    rev_headers = [
        (1, "", "FFFFFF"),
        (2, "Link do Processo", COR_VERDE_ESCURO),
        (3, "Número do processo", COR_VERDE_ESCURO),
        (4, "Nome da parte", COR_VERDE_ESCURO),
        (5, "CPF", COR_VERDE_ESCURO),
        (6, perguntas["q1"], COR_VERDE_CLARO),
        (7, perguntas["q2"], COR_VERDE_CLARO),
        (8, perguntas["q3"], COR_VERDE_CLARO),
        (9, "", "FFFFFF"),
    ]
    for col, texto, cor in rev_headers:
        cell = ws_rev.cell(2, col, texto)
        aplicar_estilo(cell, estilo_header_coluna(cor))
        if cor == "FFFFFF":
            cell.font = Font(bold=True, size=9, name="Calibri", color="000000")

    for i, caso in enumerate(casos, start=3):
        ws_rev.row_dimensions[i].height = 40
        dados = ["", caso.get("link",""), caso.get("numero_processo",""),
                 caso.get("nome_parte",""), caso.get("cpf",""), "", "", "", ""]
        for col_idx, valor in enumerate(dados, 1):
            cell = ws_rev.cell(i, col_idx, valor)
            if col_idx == 2 and valor.startswith("http"):
                aplicar_estilo(cell, estilo_link())
                cell.hyperlink = valor
            else:
                aplicar_estilo(cell, estilo_dado())

    ws_rev.freeze_panes = "A3"

    # ── ABA: Log_Revisão ───────────────────────────────────────────────
    ws_log = wb.create_sheet("Log_Revisão")
    ws_log.row_dimensions[1].height = 27.75

    log_headers = [
        "Data/Hora", "Aba Editada", "Célula", "Coluna",
        "Valor Anterior", "Valor Novo", "Inserido por", "Alterado por",
        "Tipo de Alteração", "Critério", "Squad:", "Growth",
        "Critérios Monitorados", "Subcritérios Monitorados",
        "MÉTRICAS DE REVISÃO", ""
    ]
    for col_idx, h in enumerate(log_headers, 1):
        cell = ws_log.cell(1, col_idx, h)
        aplicar_estilo(cell, estilo_header_coluna(COR_VERDE_ESCURO))
        ws_log.column_dimensions[get_column_letter(col_idx)].width = 20

    # Métricas pré-populadas
    metricas = [
        (2, 13, "F, G, H"), (2, 15, "Total de células revisadas"), (2, 16, ""),
        (3, 15, "Total de linhas revisadas"), (3, 16, ""),
        (4, 15, "Desacordos"), (4, 16, ""),
        (5, 15, "Concordâncias"), (5, 16, ""),
        (6, 15, "% Concordância"), (6, 16, ""),
        (7, 11, "Growth"), (7, 15, "Total revisado"), (7, 16, ""),
    ]
    for row, col, val in metricas:
        ws_log.cell(row, col, val)

    # ── ABA: Dados da Execução ─────────────────────────────────────────
    ws_dados = wb.create_sheet("Dados da Execução")
    ws_dados.column_dimensions["A"].width = 30
    ws_dados.column_dimensions["B"].width = 20
    ws_dados.column_dimensions["C"].width = 20

    meta = [
        ("Agente:", nome_agente),
        ("Data da rodada:", data_rodada),
        ("Total de casos:", len(casos)),
        ("Analistas:", "Mariana, Ludmila"),
        ("Data de geração:", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for row_idx, (campo, valor) in enumerate(meta, 1):
        aplicar_estilo(ws_dados.cell(row_idx, 1, campo),
                       estilo_header_coluna(COR_VERDE_ESCURO))
        ws_dados.cell(row_idx, 2, valor)

    # ── ABA: Comparação com o Agente ───────────────────────────────────
    ws_comp = wb.create_sheet("Comparação com o Agente")
    ws_comp.row_dimensions[1].height = 27.75
    ws_comp.row_dimensions[2].height = 27.75

    comp_larguras = {
        "A": 8, "B": 21.38, "C": 35.0, "D": 35.0, "E": 35.0,
        "F": 15.0, "G": 35.0, "H": 35.0, "I": 15.0, "J": 30.0
    }
    for col, w in comp_larguras.items():
        ws_comp.column_dimensions[col].width = w

    # Linha 1
    mesclar_e_estilizar(ws_comp, 1, 1, 2, "", sec_branco)
    mesclar_e_estilizar(ws_comp, 1, 3, 6, f"O PROCESSO {perguntas['q1'].upper()}", sec_verde_c)
    mesclar_e_estilizar(ws_comp, 1, 7, 7, "MODERAÇÃO", sec_verde_c)
    mesclar_e_estilizar(ws_comp, 1, 8, 9, "EM QUAL TEMA O PROCESSO SE ENQUADRA?", sec_verde_c)
    mesclar_e_estilizar(ws_comp, 1, 10, 10, "CLASSIFICAÇÃO", sec_verde_c)

    comp_h2 = [
        (1, "ID", "FFFFFF"), (2, "LINK", "FFFFFF"),
        (3, "MODERAÇÃO OQJ", COR_VERDE_CLARO),
        (4, "AGENTE (EXPANDIDO)", COR_VERDE_CLARO),
        (5, "MODERAÇÃO AGENTE", COR_VERDE_CLARO),
        (6, "DESEMPENHO DO AGENTE", COR_VERDE_CLARO),
        (7, "CLASSIFICAÇÃO OQJ", COR_VERDE_CLARO),
        (8, "CLASSIFICAÇÃO AGENTE (EXPANDIDO)", COR_VERDE_CLARO),
        (9, "DESEMPENHO DO AGENTE", COR_VERDE_CLARO),
        (10, "COMENTÁRIOS", COR_VERDE_CLARO),
    ]
    for col, texto, cor in comp_h2:
        cell = ws_comp.cell(2, col, texto)
        aplicar_estilo(cell, estilo_header_coluna(cor))
        if cor == "FFFFFF":
            cell.font = Font(bold=True, size=9, name="Calibri", color="000000")

    for i, caso in enumerate(casos, start=3):
        ws_comp.row_dimensions[i].height = 40
        cell_id = ws_comp.cell(i, 1, i - 2)
        aplicar_estilo(cell_id, estilo_dado())
        cell_link = ws_comp.cell(i, 2, caso.get("link", ""))
        if caso.get("link", "").startswith("http"):
            aplicar_estilo(cell_link, estilo_link())
            cell_link.hyperlink = caso.get("link", "")
        else:
            aplicar_estilo(cell_link, estilo_dado())
        for col in range(3, 11):
            aplicar_estilo(ws_comp.cell(i, col, ""), estilo_dado())

    ws_comp.freeze_panes = "A3"

    # ── ABA: Consulta do Desempate ─────────────────────────────────────
    ws_desemp = wb.create_sheet("Consulta do Desempate")
    # Mesma estrutura que Execução (cópia das fórmulas de desempate)
    ws_desemp.title = "Consulta do Desempate"

    ws_desemp.row_dimensions[1].height = 27.75
    ws_desemp.row_dimensions[2].height = 27.75
    ws_desemp.row_dimensions[3].height = 20

    for col, w in larguras_mae.items():
        ws_desemp.column_dimensions[col].width = w

    # Linha 1 - mesma estrutura que Execução
    mesclar_e_estilizar(ws_desemp, 1, 1, 2, "DADOS DO REPORTE", sec_branco)
    ws_desemp.cell(1, 1).font = Font(bold=True, size=10, name="Calibri", color="000000")
    mesclar_e_estilizar(ws_desemp, 1, 3, 4, "DADOS DO PROCESSO", sec_verde_e)
    mesclar_e_estilizar(ws_desemp, 1, 5, 6, "DADOS DA PARTE", sec_verde_e)
    mesclar_e_estilizar(ws_desemp, 1, 7, 14, "EXECUÇÃO", sec_verde_c)
    mesclar_e_estilizar(ws_desemp, 1, 15, 17, "DESEMPATE", sec_vermelho)

    # Linha 2
    for col in [1, 2]:
        cell = ws_desemp.cell(2, col, "")
        aplicar_estilo(cell, estilo_header_coluna("FFFFFF"))
        cell.font = Font(bold=True, size=9, name="Calibri", color="000000")
    for col, texto in [(3, "Link do Processo"), (4, "Número do processo"),
                       (5, "Nome da parte"), (6, "CPF")]:
        aplicar_estilo(ws_desemp.cell(2, col, texto), col_verde_e)
    mesclar_e_estilizar(ws_desemp, 2, 7, 8, perguntas["q1"], col_verde_c)
    mesclar_e_estilizar(ws_desemp, 2, 9, 10, perguntas["q2"], col_verde_c)
    mesclar_e_estilizar(ws_desemp, 2, 11, 12, perguntas["q3"], col_verde_c)
    aplicar_estilo(ws_desemp.cell(2, 13, "Comentários"), col_verde_c)
    aplicar_estilo(ws_desemp.cell(2, 14, "Comentários"), col_verde_c)
    ws_desemp.merge_cells("O2:O3")
    ws_desemp.merge_cells("P2:P3")
    ws_desemp.merge_cells("Q2:Q3")
    for col, texto in [(15, perguntas["q1"]), (16, perguntas["q2"]), (17, perguntas["q3"])]:
        aplicar_estilo(ws_desemp.cell(2, col, texto), col_verde_c)

    # Linha 3 nomes
    for col in [1, 2]:
        aplicar_estilo(ws_desemp.cell(3, col, "ID"), estilo_dado())
    for col in [3, 4, 5, 6]:
        aplicar_estilo(ws_desemp.cell(3, col, ""), col_verde_e)
    for col, nome, cor in [(7,"Mari",cor_mari),(8,"Lud",cor_lud),(9,"Mari",cor_mari),(10,"Lud",cor_lud),
                            (11,"Mari",cor_mari),(12,"Lud",cor_lud),(13,"Mari",cor_mari),(14,"Lud",cor_lud)]:
        aplicar_estilo(ws_desemp.cell(3, col, nome), estilo_header_analista(cor))

    # Dados
    for i, caso in enumerate(casos, start=4):
        ws_desemp.row_dimensions[i].height = 40
        dados = [caso.get("report_id",""), caso.get("lawsuit_id",""), caso.get("link",""),
                 caso.get("numero_processo",""), caso.get("nome_parte",""), caso.get("cpf","")]
        for col_idx, valor in enumerate(dados, 1):
            cell = ws_desemp.cell(i, col_idx, valor)
            if col_idx == 3 and valor.startswith("http"):
                aplicar_estilo(cell, estilo_link())
                cell.hyperlink = valor
            else:
                aplicar_estilo(cell, estilo_dado())
        for col in range(7, 18):
            aplicar_estilo(ws_desemp.cell(i, col, ""), estilo_dado())

    ws_desemp.freeze_panes = "A4"

    # ── SALVAR ─────────────────────────────────────────────────────────
    filename = f"[PLANILHA MÃE] QA - {nome_agente} {data_rodada}.xlsx"
    out = os.path.join(pasta_saida, filename)
    wb.save(out)
    print(f"  ✅ Gerado: {filename}")
    return out

# ─────────────────────────────────────────────
# SCRIPT PRINCIPAL
# ─────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  GERADOR DE PLANILHAS QA - TRUST & SAFETY (JUSBRASIL)")
    print("=" * 60)

    # 1. Arquivo de entrada
    if len(sys.argv) > 1:
        caminho_entrada = sys.argv[1]
    else:
        caminho_entrada = input("\n📂 Caminho do arquivo de entrada (.xlsx ou .csv): ").strip().strip('"')

    if not os.path.exists(caminho_entrada):
        print(f"\n❌ Arquivo não encontrado: {caminho_entrada}")
        sys.exit(1)

    # 2. Nome do agente
    print("\n🤖 nome do agente/tema da rodada")
    print("   Exemplos: Agente de Direito de Família | Agente de Saúde | Agente de Trabalho")
    nome_agente = input("   → ").strip() or "Agente"

    # 3. Data da rodada
    data_default = datetime.now().strftime("%Y.%m.%d")
    data_entrada = input(f"\n📅 Data da rodada [{data_default}]: ").strip()
    data_rodada = data_entrada or data_default

    # 4. Perguntas (usa padrão de família ou customiza)
    print("\n❓ Perguntas do formulário de execução:")
    print("   [1] Padrão Direito de Família")
    print("   [2] Padrão Saúde")
    print("   [3] Personalizado")
    opcao_q = input("   Escolha [1/2/3]: ").strip()

    if opcao_q == "2":
        perguntas = PERGUNTAS["saude"]
    elif opcao_q == "3":
        perguntas = {
            "q1": input("   Pergunta 1 (trata de X?): ").strip(),
            "q2": input("   Pergunta 2 (se enquadra?): ").strip(),
            "q3": input("   Pergunta 3 (tema?): ").strip(),
        }
    else:
        perguntas = PERGUNTAS["familia"]

    # 5. Pasta de saída
    pasta_saida = os.path.dirname(os.path.abspath(__file__))
    print(f"\n📁 Arquivos serão gerados em: {pasta_saida}")

    # 6. Ler casos
    print("\n⏳ Lendo arquivo de entrada...")
    try:
        casos = ler_arquivo_entrada(caminho_entrada)
    except Exception as e:
        print(f"\n❌ Erro ao ler arquivo: {e}")
        sys.exit(1)

    print(f"   → {len(casos)} casos encontrados.")

    if not casos:
        print("❌ Nenhum caso encontrado no arquivo. Verifique o formato.")
        sys.exit(1)

    # 7. Gerar planilhas
    print("\n⚙️  Gerando planilhas...\n")
    gerar_planilha_individual(casos, "Ludmila", perguntas, nome_agente, data_rodada, pasta_saida)
    gerar_planilha_individual(casos, "Mariana", perguntas, nome_agente, data_rodada, pasta_saida)
    gerar_planilha_mae(casos, perguntas, nome_agente, data_rodada, pasta_saida)

    print(f"\n✅ Concluído! 3 planilhas geradas em:\n   {pasta_saida}")
    print("\n   Próximos passos:")
    print("   1. Envie a planilha de Ludmila e Mariana para cada analista preencher")
    print("   2. A Planilha Mãe já está estruturada para receber as respostas")

if __name__ == "__main__":
    main()
